from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# Header + 80 data rows, A–Z. Cell count must exceed MAX_CELLS (2000).
EXPORT_DATA_ROWS = 80
EXPORT_COLS = 26
EXPORT_CELLS = (EXPORT_DATA_ROWS + 1) * EXPORT_COLS


def main() -> None:
    workbook = Workbook()

    pnl = workbook.active
    pnl.title = "P&L"
    pnl["A1"] = "Line"
    pnl["B1"] = "FY24"
    pnl["A2"] = "Revenue"
    pnl["B2"] = 1000  # hardcoded; should be Assumptions!B2*B3 (12*100 = 1200)
    pnl["A3"] = "COGS"
    pnl["B3"] = 400
    pnl["A4"] = "Gross Profit"
    pnl["B4"] = 500  # does not foot: 1000-400 = 600
    pnl["A5"] = "Opex"
    pnl["B5"] = 150
    pnl["A6"] = "Operating Profit"
    pnl["B6"] = 350  # follows the wrong GP (500-150)
    pnl["A8"] = "Notes"
    pnl["B8"] = "Revenue is not linked to Assumptions. GP/OP are hardcoded."
    pnl["D1"] = "FY25"
    pnl["E1"] = "FY26"
    pnl["F1"] = "FY27"
    # D2:F6 left empty on purpose — forecast scenario writes formulas here

    assumptions = workbook.create_sheet("Assumptions")
    assumptions["A1"] = "Driver"
    assumptions["B1"] = "Value"
    assumptions["A2"] = "Price"
    assumptions["B2"] = 12
    assumptions["A3"] = "Units"
    assumptions["B3"] = 100
    assumptions["A4"] = "YoY growth"
    assumptions["B4"] = 0.05
    assumptions["A6"] = "Note"
    assumptions["B6"] = "Price * Units should drive P&L Revenue."

    data = workbook.create_sheet("Exports")
    data["A1"] = "Month"
    data["B1"] = "Region"
    data["C1"] = "Amount"
    for col in range(4, EXPORT_COLS + 1):
        data.cell(1, col, f"Col{get_column_letter(col)}")
    regions = ["Nordics", "UK", "DACH", "US"]
    for i in range(2, EXPORT_DATA_ROWS + 2):
        data.cell(i, 1, f"2024-{((i - 2) % 12) + 1:02d}")
        data.cell(i, 2, regions[(i - 2) % 4])
        for col in range(3, EXPORT_COLS + 1):
            data.cell(i, col, 80 + ((i * 17 + col) % 40))
    if EXPORT_CELLS <= 2000:
        raise RuntimeError(
            f"Exports used range is {EXPORT_CELLS} cells; must exceed MAX_CELLS=2000"
        )

    out = Path(__file__).resolve().parents[2] / "samples" / "demo.xlsx"
    out.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(out)
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
