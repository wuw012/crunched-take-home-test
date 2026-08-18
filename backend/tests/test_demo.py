import importlib.util
from pathlib import Path

from openpyxl import load_workbook

from app.limits import MAX_CELLS


def _export_cells() -> int:
    script = Path(__file__).resolve().parents[1] / "scripts" / "make_demo.py"
    spec = importlib.util.spec_from_file_location("make_demo", script)
    assert spec is not None and spec.loader is not None
    module = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(module)
    return int(module.EXPORT_CELLS)


def test_exports_layout_exceeds_cell_cap() -> None:
    assert _export_cells() > MAX_CELLS


def test_demo_xlsx_exports_exceeds_cell_cap() -> None:
    path = Path(__file__).resolve().parents[2] / "samples" / "demo.xlsx"
    workbook = load_workbook(path, read_only=True)
    exports = workbook["Exports"]
    rows = exports.max_row or 0
    columns = exports.max_column or 0
    workbook.close()
    assert rows * columns > MAX_CELLS
